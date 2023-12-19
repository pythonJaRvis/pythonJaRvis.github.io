import os
import json
def run(truth , pycg , pythoncg):
    if isinstance(truth,int) or isinstance(pycg,int) or isinstance(pycg,int):
        return 
    TP , FN , FP  = 0 , 0 , 0
    with open(truth , 'r') as f:
        truthJson:dict = json.load(f)
    with open(pycg , 'r') as f:
        pycgJson:dict = json.load(f)
    with open(pythoncg , 'r') as f:
        pythoncgJson:dict = json.load(f)
    pycg_res = getResult(truthJson , pycgJson)
    python_res = getResult(truthJson , pythoncgJson)
    if python_res[1] or python_res[2]:
        print(pythoncg)
    # if pycg_res[1] or pycg_res[2]:
        # print(pycg)
    return pycg_res , python_res
def getEdges(cg:dict):
    cnt = 0
    for k , vList in cg.items():
        cnt += len(vList)
    return cnt
def getTP(truthJson , curJson):
    cnt = 0
    for k , vList in curJson.items():
        if k in truthJson:
            cnt += len(set(vList) & set(truthJson[k]))
    return cnt

# 假阴 多边 
def getFN(truthJson , curJson , TP):
    return getEdges(truthJson) - TP

# 假阳 错边 
def getFP(truthJson , curJson , TP):
    return getEdges(curJson) - TP

def getResult(truthJson:dict , curJson:dict):
    tp = getTP(truthJson , curJson)
    edges = getEdges(truthJson)
    FP , FN = getFP(truthJson,curJson , tp) , getFN(truthJson , curJson , tp)
    return [tp,FP,FN,edges]
    # return "{},{},{}/{}".format(tp,FN,FP,edges)
def getTime(pycgPath,pythonPath):
    pycgmem = 0
    pythonmem = 0
    with open(pycgPath,'r') as f:
        lines = f.read().split('\n')
    for line in lines:
        line = line.strip()
        # print(line)
        if line.endswith("maximum resident set size"):
            res = line.replace("maximum resident set size",'')
            res = int(res)
            pycgmem = res
    with open(pythonPath,'r') as f:
        lines = f.read().split('\n')
    for line in lines:
        line = line.strip()
        # print(line)
        if line.endswith("maximum resident set size"):
            res = line.replace("maximum resident set size",'')
            res = int(res)
            pythonmem = res      
            break
    return pycgmem,pythonmem
def findFile(base):
    for root, ds, fs in os.walk(base):
        # returnList = [0] * 3
        # for f in fs:
        #     if f == "callgraph.json":
        #         fullname = os.path.join(root, f)
        #         returnList[0] = fullname
        #     if f == 'pycg.json':
        #         fullname = os.path.join(root, f)
        #         returnList[1] = fullname
        #     if f == 'pythonCG.json':
        #         fullname = os.path.join(root, f)
        #         returnList[2] = fullname
        returnList = [0] * 2
        for f in fs:
            if f == 'pycg.json':
                fullname = os.path.join(root, f)
                returnList[0] = fullname
            if f == 'pythonCG.json':
                fullname = os.path.join(root, f)
                returnList[1] = fullname
        # yield run(returnList[0] , returnList[1] , returnList[2])
        yield getTime(returnList[0,returnList[1]])
        # main(returnList[0] , returnList[1]  ,returnList[2])
def process(pycg_list , pythoncg_list):
    pre_pycg = [0,0,0,0]
    pre_python = [0,0,0,0]
    pre_pycg = list(map(lambda x:x[0] + x[1] , zip(pre_pycg,pycg_list)))
    yield pre_pycg
global_pycg = [0,0,0,0]
global_pythoncg = [0,0,0,0]
def main(index,base):
    def save_xlsx(index,name, res):
        filename = '/Users/yixuanyan/yyx/github/supplychain/YanYixuan/pythonCG/micro-benchmark/snippets/micro.xlsx'
        from openpyxl import load_workbook
        wb = load_workbook(filename=filename)
        sheet = wb['Sheet1']
        tmprow = index
        col = 3
        for j, tmp in enumerate(res[:4]):
            tmpcol = col + j
            # print(tmprow, tmpcol, tmp)
            # print(sheet.cell(tmprow,tmpcol).value,end = '\t')
            sheet.cell(tmprow, tmpcol, value=tmp)
        print()
        sheet.cell(tmprow, 1, value=name)
        col = 7
        for j, tmp in enumerate(res[4:]):
            tmpcol = col + j
            # print(sheet.cell(tmprow,tmpcol).value , end ='\t')
            sheet.cell(tmprow, tmpcol, value=tmp)
        print()
        wb.save(filename)
        pass
    global global_pycg
    global global_pythoncg
    pre_pycg = [0,0,0,0]
    pre_python = [0,0,0,0]
    for i in findFile(base):
        if not i:
            continue
        pre_pycg = list(map(lambda x:x[0] + x[1] , zip(pre_pycg,list(i[0]))))
        pre_python = list(map(lambda x:x[0] + x[1] , zip(pre_python,list(i[1]))))
        # print(pre_pycg,pre_python)
    pycg_str = "{},{},{}/{}".format(*pre_pycg)
    python_str = "{},{},{}/{}".format(*pre_python)
    res = pre_pycg + pre_python
    # save_xlsx(190+index,base.split(os.path.sep)[-1],res)
    print(base.split(os.path.sep)[-1])
    print(pycg_str , python_str)
    global_pycg = list(map(lambda x:x[0] + x[1] , zip(pre_pycg,global_pycg)))
    global_pythoncg = list(map(lambda x:x[0] + x[1] , zip(pre_python,global_pythoncg)))

entryes = [
    '/Users/yixuanyan/yyx/github/supplychain/YanYixuan/pythonCG/micro-benchmark/snippets/assignments',
    '/Users/yixuanyan/yyx/github/supplychain/YanYixuan/pythonCG/micro-benchmark/snippets/builtins',
    '/Users/yixuanyan/yyx/github/supplychain/YanYixuan/pythonCG/micro-benchmark/snippets/classes',
    '/Users/yixuanyan/yyx/github/supplychain/YanYixuan/pythonCG/micro-benchmark/snippets/decorators',
    '/Users/yixuanyan/yyx/github/supplychain/YanYixuan/pythonCG/micro-benchmark/snippets/dicts',
    '/Users/yixuanyan/yyx/github/supplychain/YanYixuan/pythonCG/micro-benchmark/snippets/direct_calls',
    '/Users/yixuanyan/yyx/github/supplychain/YanYixuan/pythonCG/micro-benchmark/snippets/exceptions',
    '/Users/yixuanyan/yyx/github/supplychain/YanYixuan/pythonCG/micro-benchmark/snippets/functions',
    '/Users/yixuanyan/yyx/github/supplychain/YanYixuan/pythonCG/micro-benchmark/snippets/generators',
    '/Users/yixuanyan/yyx/github/supplychain/YanYixuan/pythonCG/micro-benchmark/snippets/imports',
    '/Users/yixuanyan/yyx/github/supplychain/YanYixuan/pythonCG/micro-benchmark/snippets/kwargs',
    '/Users/yixuanyan/yyx/github/supplychain/YanYixuan/pythonCG/micro-benchmark/snippets/lambdas',
    '/Users/yixuanyan/yyx/github/supplychain/YanYixuan/pythonCG/micro-benchmark/snippets/lists',
    '/Users/yixuanyan/yyx/github/supplychain/YanYixuan/pythonCG/micro-benchmark/snippets/mro',
    '/Users/yixuanyan/yyx/github/supplychain/YanYixuan/pythonCG/micro-benchmark/snippets/args',
    '/Users/yixuanyan/yyx/github/supplychain/YanYixuan/pythonCG/micro-benchmark/snippets/returns',


    '/Users/yixuanyan/yyx/github/supplychain/YanYixuan/pythonCG/micro-benchmark/snippets/newCase/args',
    '/Users/yixuanyan/yyx/github/supplychain/YanYixuan/pythonCG/micro-benchmark/snippets/newCase/assign',
    '/Users/yixuanyan/yyx/github/supplychain/YanYixuan/pythonCG/micro-benchmark/snippets/newCase/calls',
    '/Users/yixuanyan/yyx/github/supplychain/YanYixuan/pythonCG/micro-benchmark/snippets/newCase/control_flow',
    '/Users/yixuanyan/yyx/github/supplychain/YanYixuan/pythonCG/micro-benchmark/snippets/newCase/import',
]
if __name__ == "__main__":
    for index,entry in enumerate(entryes):
        main(index,entry)
    print(global_pycg,global_pythoncg)
        
# complete 不包含没有调用的边
# sound    包含了所有调用的边